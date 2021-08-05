#!/usr/bin/env python3
# -*- coding:utf-8  -*-

import time
import datetime
import os
import sys
import base64
import requests
import pandas as pd

from retry import retry
from Crypto.Cipher import AES
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from tkinter import *

# from tqdm import trange

username = ''
password = ''
p_code = ''
lab = ''
performer = ''
category = ''
sub_cat = ''
test_case = ''
df_dvt = pd.DataFrame()
df_carr = pd.DataFrame()
df = pd.DataFrame()
alt_dvt = 0
alt_num = 0
dvt_num = 0
carr_num = 0
info_dict = {}
ee1 = ''
me1 = ''
factory = ''
vin_num = ''
soft_ver = ''
build_num = ''
sub_purpose = ''
req_from = ''
delivery_date = ''
reuse_per = ''
df1 = pd.DataFrame()
df2 = pd.DataFrame()
hyperlink_dict = {}
tol_test = ''
sed_test = 0
test_short = ''
doe_purpose = ''
qty = ''
test_return = ''
alt_number = ''
status = ''
sub_date = ''
product = ''
product_type = ''
hyperlink = ''
rn = ''
wb = ''
flag = 1
soup7 = ''


def app_info():
    os.system('')
    print("\033[0;33;40mProgram Name: Reliability Test Submit. Version:20210616\033[0m")
    print('If you have any problem of using this program, pls contact Jianhua He(IT Code: hejh4)\n')


class AESCipher:
    def __init__(self, secretkey: str):
        self.key = secretkey
        self.iv = secretkey[0:16]

    def encrypt(self, text):
        pad = lambda s: s + (16 - len(s) % 16) * chr(16 - len(s) % 16)
        text = pad(text).encode()
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        encrypted_text = cipher.encrypt(text)
        return base64.b64encode(encrypted_text).decode('utf-8')

    def decrypt(self, encrypted_text):
        unpad = lambda s: s[0:-s[-1]]
        encrypted_text = base64.b64decode(encrypted_text)
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        decrypted_text = cipher.decrypt(encrypted_text)
        return unpad(decrypted_text).decode('utf-8')


def en_de():
    global username, password
    secret_key = 'Re1iabL17Y5uBm17'
    file_path = './Settings/Settings.txt'
    try:
        f = open(file_path, 'r')
        lines = f.read().splitlines()
        decryted_usr = lines[0]
        decryted_pwd = lines[1]
        username = AESCipher(secret_key).decrypt(decryted_usr)
        password = AESCipher(secret_key).decrypt(decryted_pwd)
        # print(username, password)
        f.close()
    except (IOError, IndexError):
        f = open(file_path, 'w')
        username = input('Pls input your username: ')
        password = input('Pls input your password: ')
        encrypted_usr = AESCipher(secret_key).encrypt(username)
        encrypted_pwd = AESCipher(secret_key).encrypt(password)
        lines = [encrypted_usr, encrypted_pwd]
        for line in lines:
            f.writelines(line + '\n')
        f.close()
    return username, password


class ReadExcel(object):

    def __init__(self):
        self.temp = ''
        self.wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True, data_only=True)
        self.ws_s = self.wb["Submit"]
        self.ws_a = self.wb['ALT Number']

    def read_basic(self):
        global username, password, p_code, lab, category, sub_cat, test_case, info_dict, performer, ee1, me1, factory, \
            vin_num, soft_ver, build_num, sub_purpose, req_from, delivery_date, reuse_per, wb, qty, test_type, product_type
        key_list = ['product_type', 'lab', 'p_code', 'product', 'category', 'performer', 'no_b6', 'no_b7', 'ee1', 'me1',
                    'no_b10', 'factory', 'vin_num', 'soft_ver', 'build_num', 'no_b15', 'sub_purpose', 'req_from',
                    'delivery_date', 'reuse_per', 'no_20', 'test_count', 'spl_size', 'username', 'password']
        value_list = []

        for row_info in range(1, 26):
            cell = self.ws_s.cell(row_info, 2).value
            value_list.append(cell)
        info_dict = dict(zip(key_list, value_list))
        # basic info list

        product_type = self.ws_s.cell(1, 1).value
        p_code = info_dict['product']
        lab = info_dict['lab']
        performer = info_dict['performer']
        category = info_dict['category']
        ee1 = info_dict['ee1']
        me1 = info_dict['me1']
        factory = info_dict['factory']
        vin_num = info_dict['vin_num']
        soft_ver = info_dict['soft_ver']
        build_num = info_dict['build_num']
        sub_purpose = info_dict['sub_purpose']
        req_from = info_dict['req_from']
        convert_date = info_dict['delivery_date']

        if convert_date is None:
            convert_date = datetime.date.today()

        delivery_date = convert_date.strftime("%Y/%m/%d")

        # print(delivery_date)
        reuse_per = info_dict['reuse_per']

    def read_test(self):
        # start_time = time.time()
        global test_case, df_dvt, df_carr, df, alt_dvt
        mrs = self.ws_s.max_row
        dvt_dic = []
        alt_dvt = 0
        zero_submit = self.ws_s['B22'].value

        if zero_submit == 0:
            print('PLEASE SELECT AT LEAST 1 TEST CASE')
            sys.exit()
        for row_ad in range(2, mrs + 1):
            if self.ws_s.cell(row=row_ad, column=4).value == "Y":
                temp_list = []
                for col_dvt in range(4, 12):
                    cl = self.ws_s.cell(row=row_ad, column=col_dvt).value
                    temp_list.append(cl)
                dvt_dic.append(temp_list)
                alt_dvt += 1
        df_dvt = pd.DataFrame(dvt_dic)
        # dup = dict(pd.value_counts(df_dvt.iloc[:, 5]))

        print("%s ALT/DVT" ' test cases,' % alt_dvt)

        carr_dic = []

        global carr_num, tol_test
        for row_ca in range(2, mrs + 1):
            if self.ws_s.cell(row=row_ca, column=13).value == "Y":
                temp_list = []
                for col_dvt in range(13, 20):
                    cl = self.ws_s.cell(row=row_ca, column=col_dvt).value
                    temp_list.append(cl)
                carr_dic.append(temp_list)
                carr_num += 1

        df_carr = pd.DataFrame(carr_dic)
        tol_test = alt_dvt + carr_num
        if not df_carr.empty:
            df_carr.insert(1, 'a', 'None')
            for i in range(7, 1, -1):
                df_carr.rename(columns={(i - 1): i}, inplace=True)
            df_carr.rename(columns={'a': 1}, inplace=True)

        print("%s Carrier"' test cases, \n%s test cases in total will be submitted.' % (carr_num, tol_test))
        df = df_dvt.append(df_carr, ignore_index=True)

        # print(df)

    def test_seq(self, row_sq):
        self.temp = ''
        for col in range(8):
            at = df.at[row_sq, col]
            # print(at)


class AltWebsite(object):

    def __init__(self):
        global p_code, lab, performer, category, sub_cat, test_case, doe_purpose
        self.check_dict = {}
        self.pro_dict = {}
        self.session = requests.session()
        self.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                                      'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36'
                        }

        self.script = 'ctl00$MainContentPage$ctrLookUpLocalProduct$UpdatePanel_ProductLookUp|' \
                      'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch'
        self.field = ';;AjaxControlToolkit, Version=4.1.40412.0, Culture=neutral, PublicKeyToken=' \
                     '28f01b0e84b6d53e:en-US:acfc7575-cdee-46af-964f-5d85d9cdcf92:475a4ef5:effe2a26:' \
                     '1d3ed089:5546a2b:497ef277:a43b07eb:751cdd15:dfad98a5:3cf12cf1'
        self.login_url = 'http://100.65.83.171/ALTV4/Login.aspx'
        self.submit_url = 'http://100.65.83.171/ALTV4/Registration/Submit_New.aspx'

        self.lab = lab
        self.p_code = p_code
        self.cat = category
        self.perf = performer
        self.ee1 = ee1
        self.me1 = me1
        self.factory = factory
        self.vin = vin_num
        self.sw_ver = soft_ver
        self.build = build_num
        self.req = req_from
        self.del_date = delivery_date
        self.qty = qty
        self.reuse = reuse_per
        self.dict1 = {}
        self.dict2 = {}
        self.pur_d = doe_purpose
        self.sub_cat = sub_cat
        self.test = test_case
        self.pur = sub_purpose

    @retry(tries=3, delay=1, backoff=2)
    def login(self):
        global log
        url = self.login_url
        try:
            # print('retry number')
            r = self.session.get(url, headers=self.headers, timeout=3)
            soup = BeautifulSoup(r.text, 'lxml')

            s0 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
            a0 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

            data0 = {
                '__EVENTTARGET': '',
                '__EVENTARGUMENT': '',
                '__LASTFOCUS': '',
                '__VIEWSTATE': s0,
                '__EVENTVALIDATION': a0,
                'ctl00$MainContentPage$ctrLogin$UserName': username,
                'ctl00$MainContentPage$ctrLogin$Password': password,
                'ctl00$MainContentPage$ctrLogin$LoginButton': 'Submit',
            }
            log = self.session.post(url, data=data0, headers=self.headers)
            # print(log.text)
            r0 = self.session.get(self.submit_url, headers=self.headers)
            soup = BeautifulSoup(r0.text, 'lxml')
            # print(soup)

            self.s1 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
            self.a1 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')
            for d in soup.find_all('option'):
                self.dict1.update({d.text: d['value']})
        except Exception as e:
            # print('error', e)
            if 'HTTPConnectionPool' in str(e):
                print('\033[31mConnection Error, pls check your network status! Program will be ended in 2 '
                      'seconds\033[0m')
                time.sleep(2)
                sys.exit()
            else:
                self.login_check()

    def login_check(self):
        if 'Your login attempt was not successful. Please try again.' in log.text:
            print('\033[31mWrong password:\033[0m ', password)
            print('Pls input your new password here to continue')
            if os.path.exists('./Settings/Settings.txt'):
                os.remove('./Settings/Settings.txt')
            en_de()
            self.login()
        else:
            print('Unknown error')

    @retry(tries=3, delay=1, backoff=2)
    def submit_1(self):
        self.sub_cat = sub_cat
        self.test = test_case
        self.pur = sub_purpose
        data1 = {
            '__EVENTTARGET': 'ctl00$MainContentPage$ddlLab',
            '__EVENTARGUMENT': '',
            '__VIEWSTATE': self.s1,
            '__EVENTVALIDATION': self.a1,
            'ctl00$MainContentPage$ddlLab': self.dict1[self.lab]
        }
        r1 = self.session.post(self.submit_url, data=data1, headers=self.headers, timeout=3)
        soup = BeautifulSoup(r1.text, 'lxml')
        self.s2 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        self.a2 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')
        for d in soup.find_all('option'):
            self.dict1.update({d.text: d['value']})

    @retry(tries=3, delay=1, backoff=2)
    def submit_2(self):
        global flag
        if 'Accessory' in product_type:
            data2 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$rbtnProduct$1',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s2,
                '__EVENTVALIDATION': self.a2,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 1,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch': ''
            }
            r2 = self.session.post(self.submit_url, data=data2, headers=self.headers, timeout=3)
            soup = BeautifulSoup(r2.text, 'lxml')
            self.s3 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
            self.a3 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

            for d in soup.find_all('option'):
                self.dict1.update({d.text: d['value']})
            if sub_cat == 'DVT':
                flag = 2
                print('\033[31mThe test %s submitted fail, official DVT tests cannot be performed on Accessory or '
                      'Competitor\'s product.pls change it to Eval test.\033[0m' % test_short)
                return

        if 'Moto Mobile' in product_type:
            data2 = {
                'ctl00$MainContentPage$ToolkitScriptManager1': self.script,
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s2,
                '__EVENTVALIDATION': self.a2,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch': self.p_code,
                '__ASYNCPOST': 'true'

            }

            r2 = self.session.post(self.submit_url, data=data2, headers=self.headers, timeout=3)
            soup = BeautifulSoup(r2.text, 'lxml')

            # print(soup)
            if soup.find(id='MainContentPage_ctrLookUpLocalProduct_lblNoRecordsFound'):
                print('\033[31m%s was not found in the ALT database, program terminating.\033[0m' % p_code)
                time.sleep(2)
                sys.exit()

            try:
                pro_find = soup.find_all(text=re.compile(p_code))
            except Exception as e:
                # print('error info is:', e)
                pro_find = ['Competitor or Accessory products']

            pro_list = []

            for p in pro_find:
                pro_list.append(p)
            if len(pro_list) > 1:
                print('\033[31mMultiple products found, pls choose the product you want to submit test cases.\033[0m')
                table = soup.find('table', class_='ViewGrid')
                list_product = []
                for tr in table.find_all('tr')[1:]:
                    grid = tr.find_all('td')[0:5]
                    # print(type(grid))
                    for td in grid:
                        # print(td)
                        if td.getText() == 'Not Allowed':
                            break
                        else:
                            if p_code in td.getText():
                                list_product.append(td.getText())
                # print(list_product)
                dict_product = {}
                for i in range(1, len(list_product) + 1):
                    dict_product[i] = list_product[i - 1]
                # print(dict_product)
                key = int(input('Please input the product Number that you want to submit test cases:'))
                new_p_code = dict_product[key]
                print('Your selection is:', new_p_code)
                self.p_code = new_p_code

            sa = soup.find(text=re.compile("/w"))
            sa_lists = sa.split('|')
            sa_list = [a for a in sa_lists if '/w' in a]
            self.s3 = sa_list[0]
            self.a3 = sa_list[1]

            for d in soup.find_all('option'):
                self.dict1.update({d.text: d['value']})

    @retry(tries=3, delay=1, backoff=2)
    def submit_3(self):
        if 'accessory' in product_type:
            pass
        if 'Moto Mobile' in product_type:
            data3 = {
                'ctl00$MainContentPage$ToolkitScriptManager1': self.script,
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch': self.p_code,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$gridProducts$ctl02$chkAllowSubmit': 'on',
                '__EVENTTARGET': 'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s3,
                '__EVENTVALIDATION': self.a3,
                '__ASYNCPOST': 'true'
            }
            r3 = self.session.post(self.submit_url, data=data3, headers=self.headers, timeout=3)
            soup = BeautifulSoup(r3.text, 'lxml')
            sa = soup.find(text=re.compile("/w"))
            sa_lists = sa.split('|')
            sa_list = [a for a in sa_lists if '/w' in a]
            self.s4 = sa_list[0]
            self.a4 = sa_list[1]
            for d in soup.find_all('option'):
                self.dict1.update({d.text: d['value']})

    @retry(tries=3, delay=1, backoff=2)
    def submit_4(self):
        if 'Accessory' in product_type:
            pass
        if 'Moto Mobile' in product_type:
            data4 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch': self.p_code,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ctrLookUpLocalProduct$gridProducts',
                '__EVENTARGUMENT': 'Select$0',
                '__VIEWSTATE': self.s4,
                '__EVENTVALIDATION': self.a4,
            }
            r4 = self.session.post(self.submit_url, data=data4, headers=self.headers, timeout=3)
            soup = BeautifulSoup(r4.text, 'lxml')
            self.s5 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
            self.a5 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')
            for d in soup.find_all('option'):
                self.dict1.update({d.text: d['value']})
            # print(self.dict1)

    @retry(tries=3, delay=1, backoff=2)
    def submit_5(self):
        self.dict2 = {}
        self.cat = category
        if 'Accessory' in product_type:
            data5 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': '',
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTestCategory',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s3,
                '__EVENTVALIDATION': self.a3,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 1,
                'ctl00$MainContentPage$txtOthers': self.p_code,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
            }
            # print(self.p_code)

        if 'Moto Mobile' in product_type:
            data5 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTestCategory',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s5,
                '__EVENTVALIDATION': self.a5,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ctrLookUpLocalProduct$txtProductSearch': self.p_code,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
            }

        r5 = self.session.post(self.submit_url, data=data5, headers=self.headers, timeout=3)
        soup = BeautifulSoup(r5.text, 'lxml')
        # print(soup)
        self.s6 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        self.a6 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')
        subcats = soup.find(id='MainContentPage_ddlTestSubCategory').find_all('option')
        for d in subcats:
            # soup.find_all('option'):
            # print(type(d))
            self.dict2.update({d.text: d['value']})

    @retry(tries=3, delay=1, backoff=2)
    def submit_6(self):
        if 'Accessory' in product_type:
            data6 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': '',
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTestCategory',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s6,
                '__EVENTVALIDATION': self.a6,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 1,
                'ctl00$MainContentPage$txtOthers': self.p_code,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat]
            }

        if 'Moto Mobile' in product_type:
            data6 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTestCategory',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s6,
                '__EVENTVALIDATION': self.a6,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat]
            }

        r6 = self.session.post(self.submit_url, data=data6, headers=self.headers, timeout=3)
        soup = BeautifulSoup(r6.text, 'lxml')
        # print(soup)
        self.s7 = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        self.a7 = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

        testcases = soup.find(id='MainContentPage_ddlTest').find_all('option')

        for d in testcases:
            self.dict2.update({d.text: d['value']})
        # print(test_case)

        # print(self.dict2)

    @retry(IndexError, tries=3, delay=1, backoff=1)
    def submit_7(self):

        global soup7, test_case, flag

        self.test = test_case
        if 'Accessory' in product_type:
            data7 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': '',
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTest',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s7,
                '__EVENTVALIDATION': self.a7,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 1,
                'ctl00$MainContentPage$txtOthers': self.p_code,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat],
                'ctl00$MainContentPage$ddlTest': self.dict2[self.test],
            }
            try:
                r7 = self.session.post(self.submit_url, data=data7, headers=self.headers, timeout=3)
                soup7 = BeautifulSoup(r7.text, 'lxml')
                # print(soup7)

                self.s8 = soup7.select('#__VIEWSTATE')[0].attrs.get('value')
                self.a8 = soup7.select('#__EVENTVALIDATION')[0].attrs.get('value')
                for d in soup7.find_all('option'):
                    self.dict2.update({d.text: d['value']})
                test_def = soup7.find(id="MainContentPage_lblTestDefinedForProduct")

                if 'Cannot proceed' in test_def.text:
                    flag = 2
                    return

            except Exception as e:
                print(e)
                print('Step 7 Error, retrying')
        if 'Moto Mobile' in product_type:
            data7 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': 'ctl00$MainContentPage$ddlTest',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s7,
                '__EVENTVALIDATION': self.a7,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat],
                'ctl00$MainContentPage$ddlTest': self.dict2[self.test],
            }

            try:
                r7 = self.session.post(self.submit_url, data=data7, headers=self.headers, timeout=3)
                soup7 = BeautifulSoup(r7.text, 'lxml')
                # print(soup7)

                self.s8 = soup7.select('#__VIEWSTATE')[0].attrs.get('value')
                self.a8 = soup7.select('#__EVENTVALIDATION')[0].attrs.get('value')
                for d in soup7.find_all('option'):
                    self.dict2.update({d.text: d['value']})
                test_def = soup7.find(id="MainContentPage_lblTestDefinedForProduct")

                if 'Cannot proceed, the test selected has not been defined' in test_def.text:
                    flag = 2
                    print(
                        '\033[31mThe test %s has not been defined in DVT Planning, pls contact ALT Manger!\033[0m' %
                        test_short)
                    return

                vintage_list = soup7.find(id='MainContentPage_ddlVintageList').text.strip().split('\n')
                # print(vintage_list)

                if vin_num not in vintage_list:
                    print(vin_num, 'is not defined in vintage list')
                    flag = 3

            except Exception as e:
                print(e)
                print('Step 7 Error, retrying')

    def alt_checkbox(self):
        var_list = []
        var = []
        soup = soup7
        checkbox = soup.find_all('input', type='checkbox')
        checkboxes_num = len(checkbox)
        table = soup.find('table', class_='ViewGrid')

        df_c = pd.DataFrame()
        index = 0
        col = 0
        for td in table.findAll('td'):
            text = td.getText()
            df_c.at[index, col] = text.replace("\n", "").replace("\r", "").replace(" ", "")
            col += 1
            if col % 2 == 0:
                index += 1
                col = 0
        label = 0
        for r in range(df_c.shape[0]):
            if df_c.at[r, 0]:
                label += 1
        total = label + checkboxes_num

        window = Tk()
        window.title("Select ALT Functions")
        winWidth = 1200
        winHeight = 700
        # 获取屏幕分辨率
        screenWidth = window.winfo_screenwidth()
        screenHeight = window.winfo_screenheight()
        # 计算主窗口在屏幕上的坐标
        x = int((screenWidth - winWidth) / 2)
        y = int((screenHeight - winHeight) / 2)
        window.geometry("%sx%s+%s+%s" % (winWidth, winHeight, x, y))
        # 设置窗口宽高固定
        # window.resizable(0, 0)
        vbar = Scrollbar(window, orient=VERTICAL)
        canvas = Canvas(window, yscrollcommand=vbar.set)
        vbar.config(command=canvas.yview)
        vbar.pack(side=RIGHT, fill=Y)
        frame0 = Frame(canvas)
        canvas.pack(side=TOP, expand=True, fill=BOTH)
        canvas.create_window(0, 0, window=frame0, anchor='nw')
        frame1 = Frame(frame0, width=200, height=700)
        frame2 = Frame(frame0, width=200, height=700)
        frame3 = Frame(frame0, width=200, height=700)
        frame4 = Frame(frame0, width=200, height=700)
        frame5 = Frame(frame0, width=200, height=700)
        frame1.grid(row=0, column=0)
        frame2.grid(row=0, column=1)
        frame3.grid(row=0, column=2)
        frame4.grid(row=0, column=3)
        frame5.grid(row=0, column=4)

        lbl_text = StringVar()
        Label(window, width=50, textvariable=lbl_text, fg='RED').pack()

        j = 0
        col_num = total // 5
        cbs = []
        for i in range(checkboxes_num - 1):
            txt_col1 = str(df_c.at[i, 0])
            txt_col2 = str(df_c.at[i, 1])
            # print('txt_col1 is', txt_col1)
            # print('txt_col2 is', txt_col2)
            if txt_col1 != '':
                j += 2
            else:
                j += 1

            if j < (total // 5):
                frm = frame1
            elif j < (total * 2 // 5):
                frm = frame2
            elif j < (total * 3 // 5):
                frm = frame3
            elif j < (total * 4 // 5):
                frm = frame4
            elif j < (total * 5 // 5):
                frm = frame5
            var.append(IntVar())
            if txt_col1 != '':
                Label(frm, text=txt_col1, font='Arial 8 bold').pack(side=TOP, anchor='w')
                cb = Checkbutton(frm, text=txt_col2, variable=var[-1])
                cbs.append(cb)
                cb.pack(side=TOP, anchor='w')

            else:
                cb = Checkbutton(frm, text=txt_col2, variable=var[-1])
                cbs.append(cb)
                cb.pack(side=TOP, anchor='w')

            def sel_all():
                if check_all.get() == 1:
                    for c in cbs:
                        c.select()
                        c.config(state=DISABLED)

                if check_all.get() == 0:
                    for c in cbs:
                        c.deselect()
                        c.config(state=NORMAL)

        def btn_cmd():
            nonlocal var_list
            varnum = 0
            for c in range(checkboxes_num - 1):
                v = var[c].get()
                var_list.append(v)
                varnum += v
            if varnum == 0:
                lbl_text.set('Pls select ALT Functions...')
            if varnum != 0:
                lbl_text.set('')
                window.destroy()

        check_all = IntVar()
        Checkbutton(window, text='Select All', variable=check_all, command=sel_all).pack()

        Button(window, text='OK', width=15, command=btn_cmd).pack()
        window.update()
        canvas.config(scrollregion=canvas.bbox("all"))
        window.mainloop()

        check_dict = {}

        def alt_funcs_web():
            nonlocal var_list, check_dict
            var_list = var_list
            check_dict = {}
            for ctl_no, val in enumerate(var_list):
                if val == 1:
                    if ctl_no < 10:
                        s = str(ctl_no + 2).zfill(2)
                        key = 'ctl00$MainContentPage$controlProductFunction$gridProductFunction$ctl' \
                              '%s$chkFunctionList' % s
                    else:
                        key = 'ctl00$MainContentPage$controlProductFunction$gridProductFunction$ctl' \
                              '%s$chkFunctionList' % str(ctl_no + 2)

                    check_dict.update({key: 'on'})
                else:
                    continue

        alt_funcs_web()

        self.check_dict = check_dict

    def dvt_slash(self):
        self.pro_dict = {}
        soup = soup7
        pro_ans = soup.select('#MainContentPage_controlMaterialInfo_gridProtoQuestions')
        pro_counts = str(pro_ans).count('Please provide an answer')
        # print(pro_counts)
        for i in range(2, pro_counts + 2):
            s = str(i).zfill(2)
            key = 'ctl00$MainContentPage$controlMaterialInfo$gridProtoQuestions$ctl%s$txtProtoAnswer' % s
            self.pro_dict[key] = '/'
        # print(self.pro_dict)

    # @retry(tries=3, delay=1, backoff=2)
    def submit_8(self):
        if 'Accessory' in product_type:
            data8 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': '',
                '__LASTFOCUS': '',
                '__EVENTTARGET': '',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s8,
                '__EVENTVALIDATION': self.a8,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 1,
                'ctl00$MainContentPage$txtOthers': self.p_code,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat],
                'ctl00$MainContentPage$ddlTest': self.dict2[self.test],
                'ctl00$MainContentPage$txtEE1CoreID': self.ee1,
                'ctl00$MainContentPage$txtME1CoreID': self.me1,
                'ctl00$MainContentPage$ddlFactory': self.dict2[self.factory],
                'ctl00$MainContentPage$txtVintageNumber': self.vin,
                'ctl00$MainContentPage$txtSoftwareVersion': self.sw_ver,
                'ctl00$MainContentPage$txtBuildNumber': self.build,
                'ctl00$MainContentPage$ddlSubmissionPurpose': self.dict2[self.pur],
                'ctl00$MainContentPage$txtSubmission_Purpose_Details': doe_purpose,
                'ctl00$MainContentPage$ddlRequestFrom': self.dict2[self.req],
                'ctl00$MainContentPage$txtExpStartDate': self.del_date,
                'ctl00$MainContentPage$txtQuantity': qty,
                'ctl00$MainContentPage$txtMaterialReusePercentage': self.reuse

            }

        if 'Moto Mobile' in product_type:

            data8 = {
                'MainContentPage_ToolkitScriptManager1_HiddenField': self.field,
                '__LASTFOCUS': '',
                '__EVENTTARGET': '',
                '__EVENTARGUMENT': '',
                '__VIEWSTATE': self.s8,
                '__EVENTVALIDATION': self.a8,
                'ctl00$MainContentPage$ddlLab': self.dict1[self.lab],
                'ctl00$MainContentPage$rbtnProduct': 0,
                'ctl00$MainContentPage$ddlTestCategory': self.dict1[self.cat],
                'ctl00$MainContentPage$ddlTestSubCategory': self.dict2[self.sub_cat],
                'ctl00$MainContentPage$ddlTest': self.dict2[self.test],
                'ctl00$MainContentPage$txtEE1CoreID': self.ee1,
                'ctl00$MainContentPage$txtME1CoreID': self.me1,
                'ctl00$MainContentPage$ddlFactory': self.dict2[self.factory],
                'ctl00$MainContentPage$ddlVintageList': self.dict2[self.vin],
                'ctl00$MainContentPage$txtSoftwareVersion': self.sw_ver,
                'ctl00$MainContentPage$txtBuildNumber': self.build,
                'ctl00$MainContentPage$ddlSubmissionPurpose': self.dict2[self.pur],
                'ctl00$MainContentPage$txtSubmission_Purpose_Details': doe_purpose,
                'ctl00$MainContentPage$ddlRequestFrom': self.dict2[self.req],
                'ctl00$MainContentPage$txtExpStartDate': self.del_date,
                'ctl00$MainContentPage$txtQuantity': qty,
                'ctl00$MainContentPage$txtMaterialReusePercentage': self.reuse,
            }

        # print(self.test,self.cat,self.sub_cat)

        submit_dict = {'ctl00$MainContentPage$btnSubmit': 'Submit'}

        if 'ALT' in sub_cat:
            data_last = dict(data8, **self.check_dict, **submit_dict)

        else:
            data8['ctl00$MainContentPage$ddlTestPerformedBy'] = self.dict2[self.perf]
            data_last = dict(data8, **self.pro_dict, **submit_dict)

        # 是否提交测试
        r8 = self.session.post(self.submit_url, data=data_last, headers=self.headers, timeout=3)
        # print(data_last)

        soup = BeautifulSoup(r8.text, 'lxml')
        # print(soup)

    @retry(tries=3, delay=1, backoff=2)
    def return_test_number(self):
        global sed_test, df1, df2, test_return, alt_number, status, sub_date, product, hyperlink
        r = self.session.get("http://100.65.83.171/ALTV4/Registration/SubmissionModifyList.aspx", timeout=3)
        soup = BeautifulSoup(r.text, 'lxml')
        # print(soup)
        alt_table = soup.find(id="MainContentPage_gridSubmittedTest")
        for tr in alt_table.find_all('tr')[1:2]:
            cells = tr.find_all('td')
            alt_number = cells[0].get_text().replace('\r', '').replace('\n', '')
            sub_date = cells[1].get_text()
            product = cells[2].get_text()
            test_return = cells[3].get_text()
            status = cells[4].get_text()

            hyperlink = cells[0].find('a').attrs['href']

        if test_return == test_case:
            print(alt_number, test_return)
            sed_test += 1
            sub_pur = build_num + "," + doe_purpose
            alt_list1 = [vin_num, qty, test_type, test_return, alt_number, status, sub_pur]
            alt_list2 = [new_used, delivery_date, product]
            hyper_pre = 'http://100.65.83.171/ReportServer/Pages/ReportViewer.aspx?%2fgALT_V3_Reports' \
                        '%2fALTFunction&rs:Command=Render&rc:Parameters=False&Test_Seqno='
            hyperlink_dict[alt_number] = hyper_pre + alt_number

            df1_temp = pd.DataFrame([alt_list1])
            df1 = df1.append(df1_temp, ignore_index=True)

            df2_temp = pd.DataFrame([alt_list2])
            df2 = df2.append(df2_temp, ignore_index=True)

            # print('df_alt_list1 is:', df1)
            # print(df2)

        else:
            pass
            # print(test_return, test_case)
            # print('\033[31mWarning: Submission Error, please contact hejh4 for more info.\033[5m')


def test_seq(row_sq):
    for col in range(8):
        at = df.at[row_sq, col]
        # print(at)


def ordinal(n):
    if int(n) == 11 or int(n) == 12 or int(n) == 13:
        return n + 'th'
    if n[-1] == '1':
        return n + 'st'
    if n[-1] == '2':
        return n + 'nd'
    if n[-1] == '3':
        return n + 'rd'
    return n + 'th'


def fill_excel():
    global wb, hyperlink, rn
    wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True, data_only=False)
    ws_a = wb['ALT Number']
    mra = ws_a.max_row + 1

    for rn in range(1, mra + 1):
        if ws_a.cell(rn, 4).value is None:
            # print('rn is', rn)
            break
    # print('rn value is', rn)
    tol_submit = len(df1)
    ws_a.insert_rows(rn, tol_submit)
    # print(rn)

    font = Font(name='Arial',
                size=8)

    fill_green = PatternFill(fill_type='solid',
                             start_color='FF70AD47',
                             end_color='FF70AD47')
    fill_red = PatternFill(fill_type='solid',
                           start_color='FFC00000',
                           end_color='FFC00000')
    border = Border(
        top=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
        left=Side(border_style='thin', color='FF000000'))

    start_cell = 'A' + str(rn)
    end_cell = 'U' + str(rn + tol_submit)
    for rf in ws_a[start_cell:end_cell]:
        for cell in rf:
            cell.font = font
            cell.border = border

    # print('line848', df1, df2)
    for (i, row_f) in zip(range(tol_submit), range(rn, rn + tol_submit)):
        # print('how many rows:', rn, rn + tol_test)
        # print(row_f, 'row')
        ws_a.row_dimensions[row_f].height = 10
        for col in range(0, 7):
            col_a = col + 1
            # print(row_f, col, i, col_a)
            cell_val = ws_a.cell(row=row_f, column=col_a, value=df1.at[i, col])
            cell_val.alignment = Alignment(wrapText=True)
            # print('cell value is:', cell_val.value)
        for col_2 in range(18, 21):
            col_b = col_2 + 1
            col_df = col_2 - 18
            # print(row_f, col, i, col_b)
            cell_val = ws_a.cell(row=row_f, column=col_b, value=df2.at[i, col_df])
            # print('cell value is:', cell_val.value)
        hyper_unit = 'E' + str(row_f)
        ws_a[hyper_unit].hyperlink = hyperlink_dict[ws_a[hyper_unit].value]

    wb.save('0-Master_File_Reliability.xlsm')


def window_close():
    time_left = 2
    while time_left > 0:
        print('Submitting process done. This window will be closed in %s seconds:' % time_left, end='\r')
        time.sleep(1)
        time_left = time_left - 1


def program_end():
    time_left = 2
    while time_left > 0:
        print('\033[31mVintage Error, Program will be closed in %s seconds:\033[0m' % time_left, end='\r')
        time.sleep(1)
        time_left = time_left - 1


if __name__ == '__main__':
    app_info()
    en_de()
    rd = ReadExcel()
    rd.read_basic()
    rd.read_test()
    rs = AltWebsite()
    rs.login()
    rows = alt_dvt + carr_num

    for row in range(rows):
        # global sub_cat
        test_seq(row)
        ord_num = ordinal(str(row + 1))
        print('%s out of %s test cases submitting...' % (ord_num, rows))
        qty = int(df.at[row, 2])
        new_used = df.at[row, 3]
        doe_purpose = df.at[row, 7]
        test_short = df.at[row, 4]

        sub_cat = df.at[row, 1]

        if category == 'Pre SA':
            if 'CALT' in sub_cat:
                print('WRONG Selection, there was no CALT in Pre SA phase!')
            if 'C-DVT' in sub_cat:
                print('WRONG Selection, there was no C-DVT in Pre SA phase!')

        if 'ALT' in df.at[row, 1] and len(df.at[row, 1]) < 10:
            if df.at[row, 1] == 'ALT':
                test_case = 'MD ALT 2019'
                test_type = 'Formal'

            elif 'ALT Eval' in df.at[row, 1]:
                test_case = 'MD ALT 2019 - Eval'
                test_type = 'Eval'

            elif df.at[row, 1] == 'CALT':
                test_case = 'MD C-ALT 2019'
                test_type = 'ORT'

        elif 'DVT' in df.at[row, 1]:
            if df.at[row, 1] == 'DVT':
                test_case = df.at[row, 4] + ' - ' + df.at[row, 5]
                test_type = 'Formal'

            elif df.at[row, 1] == 'DVT Eval':
                test_case = df.at[row, 4] + ' Eval - ' + df.at[row, 5]
                test_type = 'Eval'

            elif df.at[row, 1] == 'C-DVT':
                test_case = df.at[row, 4] + ' C - ' + df.at[row, 5]
                test_type = 'ORT'
        else:
            test_case = df.at[row, 5]
            test_type = 'Formal'

        if 'TMO' in df.at[row, 4] or 'ATT' in df.at[row, 4]:
            category = 'Carrier Tests'
            test_case = df.at[row, 4] + ' ' + df.at[row, 5]
            test_type = 'Carrier'
            if 'TMO' in df.at[row, 4]:
                sub_cat = 'TMO'
            elif 'ATT' in df.at[row, 4]:
                sub_cat = 'ATT'

        # print(test_case)
        # print(sub_cat)
        try:
            rs.submit_1()
            rs.submit_2()
            rs.submit_3()
            rs.submit_4()
            rs.submit_5()
            rs.submit_6()
            rs.submit_7()
            if flag != 1:
                if flag == 0:
                    rs = AltWebsite()
                    rs.login()
                    rs.submit_1()
                    rs.submit_2()
                    rs.submit_3()
                    rs.submit_4()
                    rs.submit_5()
                    rs.submit_6()
                    rs.submit_7()
                elif flag == 2:
                    flag = 1
                    continue

                elif flag == 3:
                    program_end()
                    sys.exit()
            else:
                if 'ALT' in sub_cat and len(sub_cat) < 10:
                    rs.alt_checkbox()
                if 'DVT' in sub_cat:
                    try:
                        rs.dvt_slash()
                    except UnboundLocalError:
                        continue

            rs.submit_8()
        except Exception as e:
            pass

        # print('test number')

        rs.return_test_number()

    try:
        fill_excel()
    except KeyError:
        print("There was no any data will be written into the Excel.")

    window_close()
