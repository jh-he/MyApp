import time
import os
import random
import sys
import zipfile
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tkinter import *
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from colorama import init

# import line_profiler
#
# profile = line_profiler.LineProfiler()

wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True, data_only=True)
driver = ''
username = ''
password = ''
p_code = ''
lab = ''
category = ''
sub_cat = ''
test_case = ''
df_dvt = pd.DataFrame()
df_carr = pd.DataFrame()
alt_count = 0
dvt_count = 0


def read_basic():
    global username, password, p_code, lab, category, sub_cat, test_case
    ws_a = wb['Submit']
    key_list = ['no_b1', 'lab', 'p_code', 'product', 'performer', 'no_b6', 'no_b7', 'ee1', 'me1', 'no_b10', 'factory',
                'vin_num', 'soft_ver', 'build_num', 'no_b15', 'sub_purpose', 'req_from', 'dvy_date', 'reuse_per',
                'no_20', 'test_count', 'spl_size', 'username', 'password']
    value_list = []

    for row in range(1, 25):
        cell = ws_a.cell(row, 2).value
        value_list.append(cell)
    info_dict = dict(zip(key_list, value_list))
    # basic info list
    username = info_dict['username']
    password = info_dict['password']
    p_code = info_dict['p_code']
    lab = info_dict['lab']


# @profile
def read_test():
    start = time.time()
    global test_case, df_dvt, df_carr
    ws_s = wb["Submit"]
    mrs = ws_s.max_row
    dvt_dic = []
    dvt_num = 0
    zero_submit = ws_s['B21'].value

    if zero_submit == 0:
        print('\033[1;31mPLEASE SELECT AT LEAST 1 TEST CASE.\033[0m')
        sys.exit()
    for row in range(2, mrs + 1):
        if ws_s.cell(row=row, column=3).value == "Y":
            temp_list = []
            for col_dvt in range(3, 11):
                cl = ws_s.cell(row=row, column=col_dvt).value
                temp_list.append(cl)
            dvt_dic.append(temp_list)
            dvt_num += 1
    df_dvt = pd.DataFrame(dvt_dic)
    dup = dict(pd.value_counts(df_dvt.iloc[:, 5]))

    global alt_count, dvt_count
    try:
        alt_count = dup['MD ALT 2019']
        dvt_count = dvt_num - alt_count
    except KeyError:
        alt_count = 0
        dvt_count = dvt_num
    # print(df_dvt)
    if alt_count != 0 or dvt_count != 0:
        if alt_count == 1 or alt_count == 0:
            a_case = 'case'
        if alt_count > 1:
            a_case = 'cases'
        if dvt_count == 1 or dvt_count == 0:
            d_case = 'case'
        if dvt_count > 1:
            d_case = 'cases'

        print("\033[1;34m%s ALT\033[0m" ' test ' % alt_count + a_case + ' and\n' + "\033[1;34m%s DVT\033[0m"' test '
              % dvt_count + d_case + ' and')

    carr_dic = []
    carr_num = 0
    for row in range(2, mrs + 1):
        if ws_s.cell(row=row, column=12).value == "Y":
            temp_list = []
            for col_dvt in range(12, 19):
                cl = ws_s.cell(row=row, column=col_dvt).value
                temp_list.append(cl)
            carr_dic.append(temp_list)
            carr_num += 1

    df_carr = pd.DataFrame(carr_dic)
    tol_test = dvt_num + carr_num
    if not df_carr.empty:
        df_carr.insert(1, 'a', 'None')
        for i in range(7, 1, -1):
            df_carr.rename(columns={(i - 1): i}, inplace=True)
        df_carr.rename(columns={'a': 1}, inplace=True)

    if carr_num == 1 or carr_num == 0:
        c_case = 'case'
    if carr_num > 1:
        c_case = 'cases'

    print(
        "\033[1;34m%s Carrier\033[0m"' test ' % carr_num + c_case + '\nTotal %s test ' % tol_test + c_case +
        ' will be submitted.')

    df = df_dvt.append(df_carr, ignore_index=True)
    # print(df)
    # for i in range(8):
    #     at = df.at[0, i]
    #     print(at)
    end_time = time.time() - start
    print('%.5f' % end_time)


def login_check():
    global username, password
    url = 'http://reliability.mot.com/ALTV4/Login.aspx'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/87.0.4280.88 Safari/537.36'
    }
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'lxml')

    # print(soup)
    v_state = soup.select('#__VIEWSTATE')[0].attrs.get('value')
    v_stategen = soup.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
    event_valid = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

    form_data = {
        '__VIEWSTATE': v_state,
        '__VIEWSTATEGENERATOR': v_stategen,
        '__EVENTVALIDATION': event_valid,
        'ctl00$MainContentPage$ctrLogin$UserName': username,
        'ctl00$MainContentPage$ctrLogin$Password': password,
        'ctl00$MainContentPage$ctrLogin$LoginButton': 'Submit'
    }

    # 在session中发送登录请求，此后这个session里就存储了cookie
    login_page = requests.post(url, form_data, headers=headers)
    soup = BeautifulSoup(login_page.text, 'lxml')
    login_status = soup.find('span', class_='TreeViewNodeStyle')

    if login_status is None:
        sys.exit('Alt website login failed, please make sure your username and password filled correctly, '
                 'and have a stable internet connection. After that, please restart the program.')


def app_info():
    print("\033[0;33;40m Current Version:20210127\033[0m")
    print('If you have any problem of using this program, pls contact Jianhua He(IT Code: hejh4)\n')


def start_driver():
    global driver
    try:
        chrome_opt = webdriver.ChromeOptions()
        chrome_opt.add_argument('--headless')
        driver = webdriver.Chrome(executable_path='./driver/chromedriver.exe', options=chrome_opt)

    except Exception as e:
        # driver.quit()
        error_msg = str(e)
        old_version = "This version of ChromeDriver only supports"
        if old_version in error_msg:
            chrome_version = error_msg.split("Current browser version is ")[1][0:14]
            driver_version = chrome_version.rpartition('.')[0]
            print("Updating ChromeDriver version to %s, please wait a few seconds..." % driver_version)

            # Download Chrome Driver
            rs = requests.get('https://npm.taobao.org/mirrors/chromedriver/')
            soup = BeautifulSoup(rs.text, 'lxml')
            for x in soup.find_all('a'):
                target_version = x.get_text()
                if driver_version in target_version:
                    break

            driver_url = 'https://npm.taobao.org/mirrors/chromedriver/' + target_version + 'chromedriver_win32.zip'
            user_agent = [
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                'Chrome/73.0.3683.103 Safari/537.36',
                'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0',
                'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) '
                'Version/5.1 Safari/534.50',
                'User-Agent:Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11',
                'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko',
                'Mozilla/5.0 (X11; U; Linux x86_64; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Ubuntu/10.10 '
                'Chromium/10.0.642.0 Chrome/10.0.642.0 Safari/534.16',
            ]

            headers = {
                'User-Agent': random.choice(user_agent)
            }

            print('Downloading...')
            driver_zip = requests.get(driver_url, headers=headers)
            with open('./driver/chromedriver_win32.zip', 'wb') as file:
                file.write(driver_zip.content)

            # 自动解压缩zip并删除
            print('Unzipping...')
            driver_zip = './driver/chromedriver_win32.zip'
            fz = zipfile.ZipFile(driver_zip, 'r')
            for file in fz.namelist():
                fz.extractall('./driver')
                fz.close()
            if os.path.exists(driver_zip):
                os.remove(driver_zip)

            # Restart Chrome Driver
            driver = webdriver.Chrome(executable_path='./driver/chromedriver.exe', options=chrome_opt)

            print('ChromeDriver updated, program starting...')
    print('Test case Submitting...')


def login():
    global username, password
    driver.get("http://reliability.mot.com/ALTV4/ALTManager_Home.aspx")
    driver.find_element_by_id("MainContentPage_ctrLogin_UserName").send_keys(username)
    driver.find_element_by_id("MainContentPage_ctrLogin_Password").send_keys(password)
    driver.find_element_by_id("MainContentPage_ctrLogin_LoginButton").click()


'''
def pre_content():
    driver.find_element_by_xpath(
        "(.//*[normalize-space(text()) and normalize-space(.)='New Tests'])[1]/following::span[2]").click()
    Select(driver.find_element_by_id("MainContentPage_ddlLab")).select_by_visible_text(lab)
    driver.find_element_by_id("MainContentPage_btnProductLookup").click()
    driver.find_element_by_id("MainContentPage_ctrLookUpLocalProduct_txtProductSearch").send_keys(p_code)
    driver.find_element_by_id("MainContentPage_ctrLookUpLocalProduct_txtProductSearch").send_keys(Keys.ENTER)
    driver.find_element_by_link_text("Select").click()
    Select(driver.find_element_by_id("MainContentPage_ddlTestCategory")).select_by_visible_text(category)
    Select(driver.find_element_by_id("MainContentPage_ddlTestSubCategory")).select_by_visible_text(sub_cat)
    Select(driver.find_element_by_id("MainContentPage_ddlTest")).select_by_visible_text(testcase)

    # 判断是DVT测试还是ALT测试
    if sub_cat in 'DVT Eval' or category == 'Carrier Tests':
        Select(driver.find_element_by_id("MainContentPage_ddlTestPerformedBy")).select_by_visible_text(perfmby)
    # 判读是否符合自动填写slash线的测试
    # global testshort
    if testshort in test_slash_lists.keys():
        count = test_slash_lists[testshort]
        for i in range(count):
            driver.find_element_by_id(
                "MainContentPage_controlMaterialInfo_gridProtoQuestions_txtProtoAnswer_" + str(i)).send_keys("/")

'''


def fill_excel():
    ws_a = wb['ALT Number']
    mra = ws_a.max_row

    # # ws['G2'] = 'passed'
    # # wb.save('Auto-Submission.xlsm')


def stop_driver():
    global driver
    print('Submit progress done, program closing, pls open the excel for details.')
    driver.close()
    driver.quit()


if __name__ == '__main__':
    start = time.time()
    read_test()

    end = time.time()
    print('Running time: %.5f Seconds' % (end - start))
