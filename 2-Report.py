# -*- coding: utf-8 -*-
import os
import sys
import base64
from tqdm import trange
import requests
from Crypto.Cipher import AES
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from requests_ntlm import HttpNtlmAuth

ALT_number = ''
rows_length = ''
test_status = ''
result = ''
fail_notes = ''
local_link = ''
max_row = ''
ws_a = ''
wb = ''
comp_date = ''
df2_list = pd.DataFrame()
username = ''
password = ''


def app_info():
    os.system('')
    print("\033[0;33;40mProgram Name: Reliability Test Status Check. Version:20210319\033[0m")
    print('If you have any problem of using this program, pls contact Jianhua He(IT Code: hejh4)\n')


class AESCipher:
    def __init__(self, secretkey: str):
        self.key = secretkey
        self.iv = secretkey[0:16]

    def encrypt(self, text):
        pad = lambda s: s + (16 - len(s) % 16) * chr(16 - len(s) % 16)
        text = pad(text).encode()
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        encrypted_text = cipher.encrypt(text)
        return base64.b64encode(encrypted_text).decode('utf-8')

    def decrypt(self, encrypted_text):
        unpad = lambda s: s[0:-s[-1]]
        encrypted_text = base64.b64decode(encrypted_text)
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        decrypted_text = cipher.decrypt(encrypted_text)
        return unpad(decrypted_text).decode('utf-8')


def en_de():
    global username, password
    secret_key = 'Re1iabL17Y5uBm17'
    file_path = './Settings/Settings.txt'
    try:
        f = open(file_path, 'r')
        lines = f.read().splitlines()
        decryted_usr = lines[0]
        decryted_pwd = lines[1]
        username = AESCipher(secret_key).decrypt(decryted_usr)
        password = AESCipher(secret_key).decrypt(decryted_pwd)
        # print(username, password)
        f.close()
    except (IOError, IndexError):
        f = open(file_path, 'w')
        username = input('Pls input your username: ')
        password = input('Pls input your password: ')
        encrypted_usr = AESCipher(secret_key).encrypt(username)
        encrypted_pwd = AESCipher(secret_key).encrypt(password)
        lines = [encrypted_usr, encrypted_pwd]
        for line in lines:
            f.writelines(line + '\n')
        f.close()
    return username, password


def load_excel():
    global ws_a, rows_length, df2_list, max_row, wb
    wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True)
    ws_a = wb['ALT Number']
    mra = ws_a.max_row
    test_list = []
    row_list = []
    for max_row in range(1, mra):
        if ws_a.cell(max_row, 1).value is None:
            # print('max row number is', max_row - 1)
            break
    i = 0
    for row_iter in ws_a.iter_rows(min_row=2, max_row=max_row, min_col=5, max_col=8, values_only=True):
        # print(row_iter)
        i += 1
        row_list.append(row_iter)
    df_list = pd.DataFrame(row_list, columns=['ALT Number', 'Status', 'Purpose', 'Result'])

    # print(df_list)
    # df_list['Result'] = df_list['Result'].apply(lambda s: s.strip() if type(s) == str else s)
    for x in range(0, max_row - 1):
        try:
            if df_list.at[x, 'Result'].isspace():
                df_list.at[x, 'Result'] = None
            if df_list.at[x, 'Result'] is not None:
                df_list = df_list.drop(x)
        except AttributeError as a:
            continue
    # print(max_row)
    # print(df_list)
    df2_list = df_list.reset_index()
    df2_list.rename(columns={"index": "rows_num"}, inplace=True)
    df2_list['Comp_date'] = ''
    # print(df2_list)
    rows_length = len(df2_list)
    # print(rows_length)
    if rows_length == 0:
        sys.exit()
    else:
        df_list[['Fail_notes', 'Report']] = pd.DataFrame([['NA', 'NA']])
    # print(df_list)


class Login(object):

    def __init__(self):
        # 构造Session
        self.session = requests.session()
        # self.username = ''
        # self.password = ''

        self.url = 'http://100.65.83.171/ALTV4/Login.aspx'
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/87.0.4280.88 Safari/537.36'
        }
        r = requests.get(self.url)
        soup = BeautifulSoup(r.text, 'lxml')
        v_state = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        v_stategen = soup.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
        event_valid = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

        self.data = {
            '__VIEWSTATE': v_state,
            '__VIEWSTATEGENERATOR': v_stategen,
            '__EVENTVALIDATION': event_valid,
            'ctl00$MainContentPage$ctrLogin$UserName': username,
            'ctl00$MainContentPage$ctrLogin$Password': password,
            'ctl00$MainContentPage$ctrLogin$LoginButton': 'Submit'
        }

        # 在session中发送登录请求，此后这个session里就存储了cookie
        global login_page
        login_page = self.session.post(self.url, self.data, headers=self.headers)
        soup = BeautifulSoup(login_page.text, 'lxml')
        login_status = soup.find('span', class_='TreeViewNodeStyle')

        if login_status is None:
            self.login_check()

    def login_check(self):
        if 'Your login attempt was not successful. Please try again.' in login_page.text:
            print('ATTENTION! Wrong password: ', password)
            print('Pls input your new password here to continue')
            if os.path.exists('./Settings/Settings.txt'):
                os.remove('./Settings/Settings.txt')
            en_de()
            self.login()
        else:
            print('Unknown error')

    def report_download(self):
        global ALT_number, result, fail_notes, test_status, local_link, comp_date
        local_link = ''
        # ALT_number = 'WH129354'
        report_url = "http://100.65.83.171/ALTV4/Reports/ProtoCertTestReport_Functional.aspx?Test_Seqno=" + ALT_number
        rs = self.session.get(report_url, headers=self.headers)
        soup = BeautifulSoup(rs.text, 'lxml')
        # print(soup)
        url_buttons = soup(class_='Buttons', value='Download')

        session = requests.Session()

        session.auth = HttpNtlmAuth(username, password)
#
        # 下载解压测试报告
        for url_button in url_buttons:
            file_name = url_button.get('url')
            try:
                download_link = file_name.replace(' ', '%20')
                #.replace('100.65.83.147', '100.64.182.32:200')
                # print(download_link)
            except Exception as e:
                print(e)
            # print(download_link)
            doc_name = file_name.rpartition('/')[2]
            local_link = './0-Test_Report/' + str(doc_name)
            # print('Report Downloading:', doc_name)
            headers = {

            }
            r = session.get(download_link, headers=self.headers)
            # print(r.status_code, r.content)
            with open('./0-Test_Report/' + str(doc_name), 'wb') as file:
                file.write(r.content)

        if str(url_buttons) == '':
            local_link = ''

        test_status = soup.find('span', id="MainContentPage_lblStatus_Name").get_text()
        # print(test_status)

        if test_status == 'Declined':
            result = 'NA'
        elif soup.find('span', id="MainContentPage_lblTestResult"):
            result = soup.find('span', id="MainContentPage_lblTestResult").get_text()
            # print(result)
        elif soup.find('span', id="MainContentPage_lblMTT1F"):
            result = soup.find('span', id="MainContentPage_lblMTT1F").get_text()
            # print(result)
        else:
            result = ''
            # print(result)
        fail_notes = soup.find('span', id="MainContentPage_lblALTEngrTestNotes").get_text()
        if fail_notes is None:
            fail_notes = 'NA'
        # print(fail_notes)
        comp_date = soup.find('span', id="MainContentPage_lblCompletion_Date").get_text()
        if comp_date.isspace():
            comp_date = soup.find(id="MainContentPage_lblExpCompDate").get_text()
        if not comp_date.isspace():
            comp_date = datetime.strptime(comp_date, "%m/%d/%Y").strftime("%Y/%m/%d")
            # print('completed date is:', comp_date)
        df2_list.loc[x, ['Status', 'Result', 'Fail_notes', 'Report', 'comp_date']] = [test_status, result, fail_notes,
                                                                                      local_link, comp_date]


def folder_exist():
    # 检查文件夹是否存在
    path = './0-Test_Report'
    if not os.path.exists(path):
        os.makedirs(path)


if __name__ == '__main__':
    en_de()
    lg = Login()
    app_info()
    folder_exist()
    load_excel()

    print('Checking test status...')

    for x in trange(rows_length, ncols=80):
        try:
            ALT_number = df2_list.at[x, 'ALT Number']
            lg.report_download()

            # print(ALT_number)
            # lg.report_download()
        except KeyError as k:
            continue
        except TypeError as t:
            print("Empty DataFrame")
            continue
        df2_list.loc[x, ['Status', 'Result', 'Fail_notes', 'Report', 'comp_date']] = [test_status, result, fail_notes,
                                                                                      local_link, comp_date]
    # df2_list.to_csv('last_update.csv', encoding="utf_8_sig")
    print('Writing data to Excel...')
    for row in range(rows_length):
        try:
            row_index = df2_list.at[row, 'rows_num']
            # print(row_index)
            ws_a.cell(row_index + 2, 6).value = df2_list.at[row, 'Status']
            ws_a.cell(row_index + 2, 8).value = df2_list.at[row, 'Result']
            ws_a.cell(row_index + 2, 9).value = df2_list.at[row, 'Fail_notes']
            ws_a.cell(row_index + 2, 11).value = 'Open'
            # print(df2_list.at[row, 'Report'])
            ws_a.cell(row_index + 2, 11).hyperlink = df2_list.at[row, 'Report']
            ws_a.cell(row_index + 2, 12).value = df2_list.at[row, 'comp_date']
        except Exception as e:
            print(row, e)
    # print(df_list)
    wb.save('0-Master_File_Reliability.xlsm')
    print('\nAll finished, pls open Excel file for details.')
