# -*- coding: utf-8 -*-
import os
import sys
import base64
import datetime
import requests
import pandas as pd
from tqdm import trange
from Crypto.Cipher import AES
from openpyxl import load_workbook
from bs4 import BeautifulSoup

ALT_number = ''
rows_length = ''
result = ''
fail_notes = ''
mra = ''
ws_a = ''
wb = ''
wrong_num = ''
row_x = ''
assess_done = ''
df_list = pd.DataFrame()
assess_list = ''
username = ''
password = ''


def app_info():
    os.system('')
    print("\033[0;33;40mProgram Name: Reliability Test Assessment. Version:20210319\033[0m")
    print('If you have any problem of using this program, pls contact Jianhua He(IT Code: hejh4)\n')


class AESCipher:
    def __init__(self, secretkey: str):
        self.key = secretkey
        self.iv = secretkey[0:16]

    def encrypt(self, text):
        pad = lambda s: s + (16 - len(s) % 16) * chr(16 - len(s) % 16)
        text = pad(text).encode()
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        encrypted_text = cipher.encrypt(text)
        return base64.b64encode(encrypted_text).decode('utf-8')

    def decrypt(self, encrypted_text):
        unpad = lambda s: s[0:-s[-1]]
        encrypted_text = base64.b64decode(encrypted_text)
        cipher = AES.new(key=self.key.encode(), mode=AES.MODE_CBC, IV=self.iv.encode())
        decrypted_text = cipher.decrypt(encrypted_text)
        return unpad(decrypted_text).decode('utf-8')


def en_de():
    global username, password
    secret_key = 'Re1iabL17Y5uBm17'
    file_path = './Settings/Settings.txt'
    try:
        f = open(file_path, 'r')
        lines = f.read().splitlines()
        decryted_usr = lines[0]
        decryted_pwd = lines[1]
        username = AESCipher(secret_key).decrypt(decryted_usr)
        password = AESCipher(secret_key).decrypt(decryted_pwd)
        # print(username, password)
        f.close()
    except (IOError, IndexError):
        f = open(file_path, 'w')
        username = input('Pls input your username: ')
        password = input('Pls input your password: ')
        encrypted_usr = AESCipher(secret_key).encrypt(username)
        encrypted_pwd = AESCipher(secret_key).encrypt(password)
        lines = [encrypted_usr, encrypted_pwd]
        for line in lines:
            f.writelines(line + '\n')
        f.close()
    return username, password


def load_excel():
    global ws_a, rows_length, df_list, mra, wb
    wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True)
    ws_a = wb['ALT Number']
    mra = ws_a.max_row
    row_list = []
    for max_row in range(1, mra):
        if ws_a.cell(max_row, 1).value is None:
            # print('max row number is', max_row - 1)
            max_row = max_row - 1
            break

    column_name = ['ALT Number', 'Status', 'Purpose', 'Test Result', 'Failures', 'Corrective Actions', 'Test Report',
                   'Complete Date', 'Score Assessment', 'Assessment Notes', 'Sub for Assess', 'Final Result',
                   'Update Vintage', 'Assess Done']
    i = 0
    for row in ws_a.iter_rows(min_row=2, max_row=max_row, min_col=5, max_col=18, values_only=True):
        i += 1
        row_list.append(row)
    df_list = pd.DataFrame(row_list, columns=column_name)
    # print(max_row)
    for i in range(max_row - 1):
        if df_list.at[i, 'Sub for Assess'] != 'Y':
            df_list = df_list.drop(i)

    df_list = df_list.reset_index()
    df_list.rename(columns={"index": "rows_num"}, inplace=True)
    # print(df_list)
    rows_length = len(df_list)
    # print(rows_length)
    if rows_length == 0:
        sys.exit()
    else:
        pass
    # print(df_list)


def assess_lists():
    global row_x, ALT_number, assess_done, assess_list
    ALT_number = df_list.at[row_x, 'ALT Number']
    status = df_list.at[row_x, 'Status']
    test_result = df_list.at[row_x, 'Test Result']
    failures = df_list.at[row_x, 'Failures']
    c_actions = df_list.at[row_x, 'Corrective Actions']
    comp_date = df_list.at[row_x, 'Complete Date']
    score_assess = df_list.at[row_x, 'Score Assessment']
    assess_note = df_list.at[row_x, 'Assessment Notes']
    update_vintage = df_list.at[row_x, 'Update Vintage']
    final_result = df_list.at[row_x, 'Final Result']
    assess_done = df_list.at[row_x, 'Assess Done']
    assess_list = [ALT_number, status, test_result, failures, c_actions, comp_date, score_assess, assess_note,
                   update_vintage, final_result, assess_done]
    # print(assess_list)


class Login(object):

    def __init__(self):
        # 构造Session
        self.soup_u = ''
        self.href_list = ''
        self.soup_j = ''
        self.assess_url = ''
        self.session = ''
        self.session = requests.session()
        # self.username = ''
        # self.password = ''

        self.url = 'http://100.65.83.171/ALTV4/Login.aspx'
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/87.0.4280.88 Safari/537.36'
        }
        r = requests.get(self.url)
        soup = BeautifulSoup(r.text, 'lxml')
        v_state_login = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        v_stategen_login = soup.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
        event_valid_login = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

        self.data = {
            '__VIEWSTATE': v_state_login,
            '__VIEWSTATEGENERATOR': v_stategen_login,
            '__EVENTVALIDATION': event_valid_login,
            'ctl00$MainContentPage$ctrLogin$UserName': username,
            'ctl00$MainContentPage$ctrLogin$Password': password,
            'ctl00$MainContentPage$ctrLogin$LoginButton': 'Submit'
        }

        # 在session中发送登录请求，此后这个session里就存储了cookie
        global login_page
        login_page = self.session.post(self.url, self.data, headers=self.headers)
        soup = BeautifulSoup(login_page.text, 'lxml')
        login_status = soup.find('span', class_='TreeViewNodeStyle')

        if login_status is None:
            self.login_check()

    def login_check(self):
        if 'Your login attempt was not successful. Please try again.' in login_page.text:
            print('ATTENTION! Wrong password: ', password)
            print('Pls input your new password here to continue')
            if os.path.exists('./Settings/Settings.txt'):
                os.remove('./Settings/Settings.txt')
            en_de()
            self.login()
        else:
            print('Unknown error')

    def judgement(self):
        global ALT_number, result, fail_notes, wrong_num
        # ALT_number = 'WH131059'
        self.assess_url = "http://100.65.83.171/ALTV4/ProtoCert/UpdateAttempt.aspx?Test_Seqno=" + ALT_number
        rs = self.session.get(self.assess_url, headers=self.headers)
        self.soup_j = BeautifulSoup(rs.text, 'lxml')
        if self.soup_j.find('span', id="MainContentPage_lblNotDefinedNote"):
            wrong_num = 'Y'
            print(ALT_number, 'is a ALT or Eval Test, do not need attempt assessment.')
        else:
            wrong_num = 'N'

    def assessment(self):
        # session = requests.Session()
        v_state_assess = self.soup_j.select('#__VIEWSTATE')[0].attrs.get('value')
        v_stategen_assess = self.soup_j.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
        # print(v_state_assess)
        # print(v_stategen_assess)
        eng_ass_dict = {'Passed': 1, 'Failed': 5}  # cmbALTEngassessment, must
        eng_assess = eng_ass_dict[assess_list[2]]

        if 'Pass' in assess_list[2]:
            eng_notes = 'passed'  # txtALTEngJustification, must

        else:
            eng_notes = assess_list[3]
        corr_action = assess_list[4]  # txtCorrectiveAction, optional
        test_comp_date = assess_list[5]  # txtTestDate format mm/dd/yyyy, must
        sub_f_appv = 'on'  # chkWaittingForApporove, optional(不需要不传入)
        mgr_assess_dict = {'Not Started': 1, 'In Progress': 2, 'Passed': 3, 'Failed': 6}
        mgr_assess = mgr_assess_dict[assess_list[2]]
        # print('assess list 0-7 is', assess_list[0:8])
        # if 'PASS' in assess_list[2]:
        #     mgr_assess = mgr_assess_dict['Passed']  # cmbStatus, get from excel, map from mgr_assess_dict must
        # else:
        #     mgr_assess = mgr_assess_dict['Failed']
        Score_assess_dict = {'FAIL Sev 1': 7,
                             'FAIL Sev 2': 6,
                             'FAIL Sev 3': 5,
                             'FAIL Sev 4': 2,
                             'FAIL Sev 5': 3}
        mgr_notes = assess_list[3]
        apprv_by = username  # Approver Core ID, must

        # Assessment date, must
        assess_date = datetime.date.today().strftime("%m/%d/%Y")
        # print(assess_date)

        form_data_assess = {
            '__VIEWSTATE': v_state_assess,
            '__VIEWSTATEGENERATOR': v_stategen_assess,
            'ctl00$MainContentPage$cmbALTEngAssessment': eng_assess,
            'ctl00$MainContentPage$txtALTEngJustification': eng_notes,
            'ctl00$MainContentPage$txtCorrectiveAction': corr_action,
            'ctl00$MainContentPage$txtTestDate': test_comp_date,
            # 'ctl00$MainContentPage$chkWaittingForApporove': sub_f_appv,
            'ctl00$MainContentPage$cmbStatus': mgr_assess,
            'ctl00$MainContentPage$txtApprovalNote': mgr_notes,
            'ctl00$MainContentPage$txtApprovedBy': apprv_by,
            'ctl00$MainContentPage$txtApprovedDate': assess_date,
            'ctl00$MainContentPage$btnUpdate': 'Update Attempt'
        }
        form_data_assess_mgr_only = {
            '__VIEWSTATE': v_state_assess,
            '__VIEWSTATEGENERATOR': v_stategen_assess,
            'ctl00$MainContentPage$txtCorrectiveAction': corr_action,
            # 'ctl00$MainContentPage$chkWaittingForApporove': sub_f_appv,
            'ctl00$MainContentPage$cmbStatus': mgr_assess,
            'ctl00$MainContentPage$txtApprovalNote': mgr_notes,
            'ctl00$MainContentPage$txtApprovedBy': apprv_by,
            'ctl00$MainContentPage$txtApprovedDate': assess_date,
            'ctl00$MainContentPage$btnUpdate': 'Update Attempt'
        }

        assessment = self.session.post(self.assess_url, form_data_assess, headers=self.headers)

        soup = BeautifulSoup(assessment.text, 'lxml')

        href = soup.find('a', id="MainContentPage_hlnkTestStatus").get('href')
        href = href.replace('(', ',').replace("'", '')
        self.href_list = href.split(',')[1:4]
        # print(href_list)

        # 检查是否成功
        rc = self.session.get(self.assess_url, headers=self.headers)
        soup_c = BeautifulSoup(rc.text, 'lxml')
        # print(soup_c)
        # print(soup_c.find(id="MainContentPage_txtApprovedBy")['value'])
        if soup_c.find(id="MainContentPage_txtApprovedBy").has_attr('value'):
            assess_list[10] = 'Y'
        else:
            assess_list[10] = 'N'

            status_url = 'http://100.65.83.171/ALTV4/Registration/UpdateRegistration.aspx?Test_Seqno' \
                         '=' + ALT_number
            ru = self.session.get(status_url, headers=self.headers)
            self.soup_u = BeautifulSoup(ru.text, 'lxml')
            # session = requests.Session()

            v_state_status = self.soup_u.select('#__VIEWSTATE')[0].attrs.get('value')
            # print(v_state_status)
            v_stategen_status = self.soup_u.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
            # print(v_stategen_status)

            status_update_dict = {'In-Progress': 13, 'Completed': 14, 'Closed': 15}  # cmbALTEngassessment, must
            status_change = status_update_dict['In-Progress']

            form_data_status = {
                '__VIEWSTATE': v_state_status,
                '__VIEWSTATEGENERATOR': v_stategen_status,
                'ctl00$MainContentPage$ddlNewStatus': '13',
                'ctl00$MainContentPage$btnModify': 'Update'
            }

            sta_change = self.session.post(status_url, form_data_status, headers=self.headers)

            soup = BeautifulSoup(sta_change.text, 'lxml')

            print(ALT_number, ' assessed failed, pls open LINK1 to change test status, and LINK2 to add failures.')
            status_url = 'http://100.65.83.171/ALTV4/Registration/UpdateRegistration.aspx?Test_Seqno=' + ALT_number
            failure_url = 'http://100.65.83.171/ALTV4/Failure/Failure_List.aspx?Test_Seqno=' + ALT_number
            print('LINK1', status_url)
            print('LINK2', failure_url)

    def update_vin(self):
        appv_url = 'http://100.65.83.171/ALTV4/ProtoCert/UpdateProtoCertTestStatus.aspx?' \
                   'ProtoCertDefinedTestID=' + self.href_list[0] + '&ProductID=' + self.href_list[2] \
                   + '&VintageID=' + self.href_list[1] + '&TestSeqNo=' + ALT_number
        ra = self.session.get(appv_url, headers=self.headers)
        soup_a = BeautifulSoup(ra.text, 'lxml')
        # print(soup_a)

        v_state_update = soup_a.select('#__VIEWSTATE')[0].attrs.get('value')
        v_stategen_update = soup_a.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
        event_valid_update = soup_a.select('#__EVENTVALIDATION')[0].attrs.get('value')
        # print(event_valid_update)
        severity_level = {'PASS': 1, 'PASS w Re-confirm': 9, 'FAIL': 8, 'FAIL Sev 5': 2, 'FAIL Sev 4': 3,
                          'FAIL Sev 3': 5, 'FAIL Sev 2': 6, 'FAIL Sev 1': 7, 'Not Yet Evaluated': 4}

        # print(score_assessment)
        if assess_list[8] == 'Y':
            print(appv_url)
            final_result_dict = {'Y': 'on', 'N': 'off'}
            # Final Result
            if assess_list[9] == 'Y':
                final_result = final_result_dict[assess_list[9]]  # final result from excel, map from dict
            else:
                final_result = ''

            score_assessment = severity_level[assess_list[6]]  # Test result, get from excel, map severity level
            print(assess_list[6], score_assessment)

            # print(final_result)
            approve_note = assess_list[7]  # approve note from excel
            print(approve_note)
            form_data_update = {
                '__VIEWSTATE': v_state_update,
                '__VIEWSTATEGENERATOR': v_stategen_update,
                '__EVENTVALIDATION': event_valid_update,
                'ddlScoreAssessment': score_assessment,
                'txtApprovalNote': approve_note,
                'txtCoreID': username,
                'txtTest_Seqno': ALT_number,
                'chkActiveResult': final_result,
                'btnUpdate': 'Update'
            }
            app_update = self.session.post(appv_url, form_data_update, headers=self.headers)


if __name__ == '__main__':
    en_de()
    lg = Login()
    app_info()
    load_excel()

    # for i in range(rows_length):
    for i in trange(rows_length, ncols=80):
        row_x = i
        assess_lists()
        # print('outer', assess_list)
        lg.judgement()
        if wrong_num == 'Y':
            # print('if', ALT_number)
            continue
        else:
            # print('else', ALT_number)
            lg.assessment()
            df_list.at[row_x, 'Assess Done'] = assess_list[10]
            lg.update_vin()

    for x in range(rows_length):
        row_num = df_list.at[x, 'rows_num'] + 2
        # print('row number is', row_num)
        col_val = df_list.at[x, 'Assess Done']
        # print('assess done value is', col_val)
        ws_a.cell(row=row_num, column=18, value=col_val)

    wb.save('0-Master_File_Reliability.xlsm')
    print('Assessment finished, pls open Excel file for details.')
