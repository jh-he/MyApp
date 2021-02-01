import requests
from bs4 import BeautifulSoup
from requests_ntlm import HttpNtlmAuth
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import winreg





wb = load_workbook('0-Master_File_Reliability.xlsm', keep_vba=True)
ws_s = wb["Submit"]
ws_a = wb['ALT Number']

mrs = ws_s.max_row
mra = ws_a.max_row

# print(mrs, mra)

# print(ws_s['C2':'J52'])

dvt_dic = []
for rdvt in range(2, mrs + 1):
    temp_list = []
    if ws_s.cell(row=rdvt, column=3).value == "Y":
        # pid = rdvt
        cc = ws_s.cell(row=rdvt, column=3).value
        cd = ws_s.cell(row=rdvt, column=4).value
        ce = ws_s.cell(row=rdvt, column=5).value
        cf = ws_s.cell(row=rdvt, column=6).value
        cg = ws_s.cell(row=rdvt, column=7).value
        ch = ws_s.cell(row=rdvt, column=8).value
        ci = ws_s.cell(row=rdvt, column=9).value
        cj = ws_s.cell(row=rdvt, column=10).value

        temp_dvt = [cc, cd, ce, cf, cg, ch, ci, cj]
        dvt_dic.append(temp_dvt)
dvt_count = len(dvt_dic)
print("共计需要提交 %s 项DVT测试" % dvt_count)

car_dic = []
for rx in range(2, mrs + 1):
    if ws_s.cell(row=rx, column=12).value == "Y":
        cl = ws_s.cell(row=rx, column=12).value
        cm = ws_s.cell(row=rx, column=13).value
        cn = ws_s.cell(row=rx, column=14).value
        co = ws_s.cell(row=rx, column=15).value
        cp = ws_s.cell(row=rx, column=16).value
        cq = ws_s.cell(row=rx, column=17).value
        cr = ws_s.cell(row=rx, column=18).value

        temp_car = [cl, cm, cn, co, cp, cq, cr]
        car_dic.append(temp_car)
car_count = len(dvt_dic)

print("共计需要提交 %s 项Carrier测试" % dvt_count)

df_dvt = pd.DataFrame.from_dict(dvt_dic)
df_car = pd.DataFrame.from_dict(car_dic)

print(df_dvt)
print(df_car)
# ws['G2'] = 'passed'
# wb.save('Auto-Submission.xlsm')


'''
class login(object):

    def __init__(self):
        # 构造Session
        self.session = requests.session()
        self.username = 'hejh4'
        self.password = 'J1anhua!3'

        self.url = 'http://reliability.mot.com/ALTV4/Login.aspx'
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/87.0.4280.88 Safari/537.36'
        }
        r = requests.get(self.url)
        soup = BeautifulSoup(r.text, 'lxml')
        v_state = soup.select('#__VIEWSTATE')[0].attrs.get('value')
        v_stategen = soup.select('#__VIEWSTATEGENERATOR')[0].attrs.get('value')
        event_valid = soup.select('#__EVENTVALIDATION')[0].attrs.get('value')

        self.data = {
            '__VIEWSTATE': v_state,
            '__VIEWSTATEGENERATOR': v_stategen,
            '__EVENTVALIDATION': event_valid,
            'ctl00$MainContentPage$ctrLogin$UserName': self.username,
            'ctl00$MainContentPage$ctrLogin$Password': self.password,
            'ctl00$MainContentPage$ctrLogin$LoginButton': 'Submit'
        }

        # 在session中发送登录请求，此后这个session里就存储了cookie
        self.session.post(self.url, self.data, headers=self.headers)

    def submit(self):
        ALT_number = 'WH128516'
        report_url = "http://reliability.mot.com/ALTV4/Reports/ProtoCertTestReport_Functional.aspx?Test_Seqno=" + ALT_number
        rs = self.session.get(report_url, headers=self.headers)
        soup = BeautifulSoup(rs.text, 'lxml')
        url_buttons = soup(class_='Buttons', value='Download')
        session = requests.Session()
        session.auth = HttpNtlmAuth(self.username, self.password)
        # session.get('http://100.64.182.32:200/')
        for url_button in url_buttons:
            download_link = url_button.get('url')
            doc_name = download_link.rpartition('/')[2]
            r = session.get(download_link)
            print(r.status_code)
            with open('./Test Report/' + str(doc_name), 'wb') as file:
                file.write((r.content))


if __name__ == '__main__':
    login().submit()
'''
